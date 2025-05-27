import openpyxl

file_path = "/Volumes/KINGSTON/Software Testing/data_driven_testing.xlsx"

workbook = openpyxl.load_workbook(file_path)
# sheet = workbook["Data"]
sheet = workbook.active  # get active sheet from excel

for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(r, c).value = "Welcome"

workbook.save(file_path)

# multiple data
file_path = "/Volumes/KINGSTON/Software Testing/test_data.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

sheet.cell(1, 1).value = "Employee ID"
sheet.cell(1, 2).value = "Name"
sheet.cell(1, 3).value = "Position"

sheet.cell(2, 1).value = 123
sheet.cell(2, 2).value = "Smith"
sheet.cell(2, 3).value = "Engineer"

sheet.cell(3, 1).value = 321
sheet.cell(3, 2).value = "David"
sheet.cell(3, 3).value = "Manager"

sheet.cell(4, 1).value = 456
sheet.cell(4, 2).value = "John"
sheet.cell(4, 3).value = "Software Tester"

workbook.save(file_path)  # save file after entering the data
