import openpyxl

# file > workbook > sheet > row > cell
file_path = "/Volumes/KINGSTON/Software Testing/data_driven_testing.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Sheet1"]
num_row = sheet.max_row  # count the number of rows
num_column = sheet.max_column  # count the number of columns

for r in range(1, num_row + 1):
    for c in range(1, num_column + 1):
        print(sheet.cell(r, c).value, end=' | ')
    print()
