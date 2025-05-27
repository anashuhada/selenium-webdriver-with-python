import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file_path, sheet_name, num_row, num_column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(num_row, num_column).value


def write_data(file_path, sheet_name, num_row, num_column, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(num_row, num_column).value = data
    workbook.save(file_path)


def fill_green_color(file_path, sheet_name, num_row, num_column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(start_color="60d212", end_color="60d212", fill_type="solid")
    sheet.cell(num_row, num_column).fill = green_fill
    workbook.save(file_path)


def fill_red_color(file_path, sheet_name, num_row, num_column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
    sheet.cell(num_row, num_column).fill = red_fill
    workbook.save(file_path)